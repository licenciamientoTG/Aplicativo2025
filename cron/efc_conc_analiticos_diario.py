"""Importación diaria de Analíticos REGIO, completamente en Python.

Lee IMAP en modo sólo lectura, analiza PLANILLA y escribe únicamente tablas
TG.dbo.efc_conc_*. No usa PHP, SG12 ni movimientos_bancarios.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime
import email
from email.header import decode_header
from email.utils import parsedate_to_datetime
import hashlib
from io import BytesIO
import imaplib
import json
import os
from pathlib import Path
import sys
import unicodedata

import pyodbc
import xlrd
from openpyxl import load_workbook


VERSION = "2.0-python"
SCRIPT_DIR = Path(__file__).resolve().parent
ROOT = SCRIPT_DIR.parent


def load_env_file() -> None:
    """Carga .env junto al script, desde la carpeta actual o desde la raíz."""
    paths = (SCRIPT_DIR / ".env", Path.cwd() / ".env", ROOT / ".env")
    path = next((candidate for candidate in paths if candidate.is_file()), None)
    if path is None:
        return
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, value = line.split("=", 1)
        os.environ.setdefault(name.strip(), value.strip().strip('"').strip("'"))


def decode(value: str | None) -> str:
    return "".join(piece.decode(charset or "utf-8", errors="replace") if isinstance(piece, bytes) else piece for piece, charset in decode_header(value or ""))


def key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").upper()
    return "".join(char for char in text if char.isalnum())


def digits(value: object) -> str:
    return "".join(char for char in str(value or "") if char.isdigit()).lstrip("0")


def money(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "").replace("$", "").strip()
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def as_date(value: object, datemode: int | None = None) -> date | None:
    if isinstance(value, datetime):
        return value.date() if value.year >= 2020 else None
    if isinstance(value, date):
        return value if value.year >= 2020 else None
    if isinstance(value, (int, float)) and datemode is not None and 30000 < value < 80000:
        parsed = xlrd.xldate.xldate_as_datetime(value, datemode).date()
        return parsed if parsed.year >= 2020 else None
    if value:
        for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%A, %B %d, %Y"):
            try:
                parsed = datetime.strptime(str(value).strip(), pattern).date()
                return parsed if parsed.year >= 2020 else None
            except ValueError:
                pass
    return None


def db_connection() -> pyodbc.Connection:
    required = ("EFC_CONC_DB_HOST", "EFC_CONC_DB_NAME", "EFC_CONC_DB_USER", "EFC_CONC_DB_PASSWORD")
    missing = [name for name in required if not os.environ.get(name, "").strip()]
    if missing:
        raise RuntimeError("Faltan variables de conexión TG: " + ", ".join(missing))
    host = os.environ["EFC_CONC_DB_HOST"].strip()
    port = os.environ.get("EFC_CONC_DB_PORT", "1433").strip()
    driver = os.environ.get("EFC_CONC_DB_DRIVER", "ODBC Driver 17 for SQL Server").strip()
    trust = os.environ.get("EFC_CONC_DB_TRUST_CERTIFICATE", "yes").strip()
    connection_string = f"DRIVER={{{driver}}};SERVER={host},{port};DATABASE={os.environ['EFC_CONC_DB_NAME'].strip()};UID={os.environ['EFC_CONC_DB_USER'].strip()};PWD={os.environ['EFC_CONC_DB_PASSWORD']};TrustServerCertificate={trust};"
    return pyodbc.connect(connection_string, autocommit=False)


def ensure_schema(cursor: pyodbc.Cursor) -> None:
    statements = [
        """IF OBJECT_ID('dbo.efc_conc_analiticos_importaciones','U') IS NULL
            CREATE TABLE dbo.efc_conc_analiticos_importaciones (
                id INT IDENTITY PRIMARY KEY, hash_archivo CHAR(64) NOT NULL,
                mensaje_uid VARCHAR(80) NULL, correo_origen NVARCHAR(320) NULL,
                asunto NVARCHAR(500) NULL, fecha_recepcion DATETIME NULL,
                nombre_archivo NVARCHAR(260) NOT NULL, mime_type VARCHAR(120) NULL,
                archivo VARBINARY(MAX) NOT NULL, version_lector VARCHAR(20) NOT NULL,
                estado VARCHAR(20) NOT NULL, total_papeletas INT NOT NULL DEFAULT 0,
                total_errores INT NOT NULL DEFAULT 0, mensaje_error NVARCHAR(MAX) NULL,
                creado_en DATETIME NOT NULL DEFAULT GETDATE(), actualizado_en DATETIME NULL,
                CONSTRAINT UQ_efc_conc_analiticos_hash UNIQUE(hash_archivo))""",
        """IF OBJECT_ID('dbo.efc_conc_analiticos_papeletas','U') IS NULL
            CREATE TABLE dbo.efc_conc_analiticos_papeletas (
                id INT IDENTITY PRIMARY KEY, importacion_id INT NOT NULL, hoja NVARCHAR(80) NOT NULL,
                fila_origen INT NOT NULL, fecha_reportada DATE NULL, fecha_original NVARCHAR(100) NULL,
                hora_original NVARCHAR(100) NULL, estacion_original NVARCHAR(250) NULL,
                estacion_nombre_original NVARCHAR(250) NULL, estacion_id INT NULL,
                estado_estacion VARCHAR(30) NOT NULL, cuenta_mn_original NVARCHAR(100) NULL,
                cuenta_mn_normalizada VARCHAR(50) NULL, remesa_numero NVARCHAR(100) NULL,
                dice_contener_mn DECIMAL(18,2) NULL, real_mn DECIMAL(18,2) NULL,
                diferencia_mn DECIMAL(18,2) NULL, dice_contener_usd DECIMAL(18,2) NULL,
                real_usd DECIMAL(18,2) NULL, diferencia_usd DECIMAL(18,2) NULL,
                datos_originales NVARCHAR(MAX) NOT NULL, creado_en DATETIME NOT NULL DEFAULT GETDATE(),
                CONSTRAINT FK_efc_conc_analiticos_papeletas_importacion FOREIGN KEY(importacion_id) REFERENCES dbo.efc_conc_analiticos_importaciones(id),
                CONSTRAINT UQ_efc_conc_analiticos_fila UNIQUE(importacion_id,hoja,fila_origen))""",
        """IF OBJECT_ID('dbo.efc_conc_analiticos_errores','U') IS NULL
            CREATE TABLE dbo.efc_conc_analiticos_errores (
                id INT IDENTITY PRIMARY KEY, importacion_id INT NOT NULL, hoja NVARCHAR(80) NULL,
                fila_origen INT NULL, tipo VARCHAR(40) NOT NULL, detalle NVARCHAR(MAX) NOT NULL,
                datos_originales NVARCHAR(MAX) NULL, creado_en DATETIME NOT NULL DEFAULT GETDATE(),
                CONSTRAINT FK_efc_conc_analiticos_errores_importacion FOREIGN KEY(importacion_id) REFERENCES dbo.efc_conc_analiticos_importaciones(id))""",
        """IF OBJECT_ID('dbo.efc_conc_analiticos_alias_estacion','U') IS NULL
            CREATE TABLE dbo.efc_conc_analiticos_alias_estacion (
                id INT IDENTITY PRIMARY KEY, alias_normalizado VARCHAR(160) NOT NULL,
                estacion_id INT NOT NULL, activo BIT NOT NULL DEFAULT 1,
                creado_en DATETIME NOT NULL DEFAULT GETDATE(), actualizado_en DATETIME NULL,
                CONSTRAINT UQ_efc_conc_analiticos_alias UNIQUE(alias_normalizado))""",
        """IF NOT EXISTS(SELECT 1 FROM sys.indexes WHERE name='IX_efc_conc_analiticos_estacion_fecha')
            CREATE INDEX IX_efc_conc_analiticos_estacion_fecha ON dbo.efc_conc_analiticos_papeletas(estacion_id,fecha_reportada)""",
    ]
    for statement in statements:
        cursor.execute(statement)
    row = cursor.execute("SELECT TOP 1 Codigo FROM TG.dbo.Estaciones WHERE RFC='DGA930823KD3' AND UPPER(Nombre) LIKE '%TRAVEL%CENTER%' ORDER BY Codigo").fetchone()
    if row:
        exists = cursor.execute("SELECT 1 FROM dbo.efc_conc_analiticos_alias_estacion WHERE alias_normalizado='KM300'").fetchone()
        if exists:
            cursor.execute("UPDATE dbo.efc_conc_analiticos_alias_estacion SET estacion_id=?,activo=1,actualizado_en=GETDATE() WHERE alias_normalizado='KM300'", row[0])
        else:
            cursor.execute("INSERT dbo.efc_conc_analiticos_alias_estacion(alias_normalizado,estacion_id) VALUES('KM300',?)", row[0])


def catalog(cursor: pyodbc.Cursor) -> tuple[list[tuple], dict[str, int]]:
    stations = cursor.execute("SELECT Codigo,Nombre,Estacion FROM TG.dbo.Estaciones WHERE RFC='DGA930823KD3'").fetchall()
    aliases = {row[0]: int(row[1]) for row in cursor.execute("SELECT alias_normalizado,estacion_id FROM dbo.efc_conc_analiticos_alias_estacion WHERE activo=1")}
    return stations, aliases


def resolve_station(raw: object, name: object, stations: list[tuple], aliases: dict[str, int]) -> tuple[int | None, str]:
    candidates = [item for item in (key(raw), key(name)) if item]
    for candidate in candidates:
        if candidate in aliases:
            return aliases[candidate], "IDENTIFICADA_ALIAS"
    hits: set[int] = set()
    for station_id, station_name, bank_code in stations:
        code, station_key = digits(bank_code), key(station_name)
        if any((code and candidate == code) or (station_key and candidate == station_key) for candidate in candidates):
            hits.add(int(station_id))
    return (next(iter(hits)), "IDENTIFICADA") if len(hits) == 1 else (None, "ESTACION_AMBIGUA" if len(hits) > 1 else "ESTACION_NO_IDENTIFICADA")


def map_headers(values: list[object]) -> dict[str, int]:
    headers: dict[str, int] = {}
    differences = 0
    for index, value in enumerate(values):
        header = key(value)
        if header == "FECHA": headers["date"] = index
        elif header == "ESTACION": headers["station"] = index
        elif "NOMBREDELAESTACIONDESERVICIO" in header: headers["station_name"] = index
        elif header == "HORA": headers["time"] = index
        elif "DECUENTAMN" in header or "NUMERODECUENTAMN" in header: headers["account"] = index
        elif "REMNUM" in header or "REMESA" in header: headers["remittance"] = index
        elif "DICECONTENERMN" in header: headers["declared_mn"] = index
        elif "REALMN" in header: headers["real_mn"] = index
        elif "DICECONTENERDLL" in header or "DICECONTENERDLS" in header: headers["declared_usd"] = index
        elif "REALDLS" in header or "REALDLL" in header: headers["real_usd"] = index
        elif header == "DIFERENCIA":
            differences += 1
            headers["difference_mn" if differences == 1 else "difference_usd"] = index
    return headers


def workbook_rows(content: bytes, filename: str) -> tuple[list[list[object]], int | None]:
    if filename.lower().endswith(".xlsx"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        if "PLANILLA" not in workbook.sheetnames:
            raise RuntimeError("El archivo no contiene la hoja PLANILLA.")
        return [list(row) for row in workbook["PLANILLA"].iter_rows(values_only=True)], None
    workbook = xlrd.open_workbook(file_contents=content)
    try:
        sheet = workbook.sheet_by_name("PLANILLA")
    except xlrd.biffh.XLRDError as exc:
        raise RuntimeError("El archivo no contiene la hoja PLANILLA.") from exc
    return [sheet.row_values(row) for row in range(sheet.nrows)], workbook.datemode


def parse_workbook(content: bytes, filename: str, stations: list[tuple], aliases: dict[str, int]) -> tuple[list[tuple], list[tuple]]:
    rows, datemode = workbook_rows(content, filename)
    header_index, headers = next(((index, mapped) for index, row in enumerate(rows[:60]) if {"station", "declared_mn", "real_mn"}.issubset(mapped := map_headers(row))), (None, {}))
    if header_index is None:
        raise RuntimeError("No se encontraron los encabezados requeridos en PLANILLA.")
    report_date = next((parsed for row in rows[:8] for parsed in (as_date(value, datemode) for value in row) if parsed), None)
    if not report_date:
        try: report_date = datetime.strptime(filename[-14:-4], "%d-%m-%Y").date()
        except ValueError: pass
    output, errors = [], []
    for source_row, row in enumerate(rows[header_index + 1:], header_index + 2):
        value = lambda field: row[headers[field]] if field in headers and headers[field] < len(row) else None
        declared_mn, real_mn, declared_usd, real_usd = money(value("declared_mn")), money(value("real_mn")), money(value("declared_usd")), money(value("real_usd"))
        if not any(item is not None for item in (declared_mn, real_mn, declared_usd, real_usd)):
            continue
        paper_date = as_date(value("date"), datemode) or report_date
        raw_json = json.dumps(row, ensure_ascii=False, default=str)
        if not paper_date:
            errors.append((source_row, "FECHA_NO_IDENTIFICADA", "No fue posible identificar la fecha de la papeleta.", raw_json))
            continue
        station_id, station_status = resolve_station(value("station"), value("station_name"), stations, aliases)
        account = str(value("account") or "").strip() or None
        output.append((source_row, paper_date.isoformat(), str(value("date") or "") or None, str(value("time") or "") or None, str(value("station") or "") or None, str(value("station_name") or "") or None, station_id, station_status, account, "".join(char for char in account or "" if char.isdigit()) or None, str(value("remittance") or "") or None, declared_mn, real_mn, money(value("difference_mn")), declared_usd, real_usd, money(value("difference_usd")), raw_json))
    if not output:
        raise RuntimeError("PLANILLA no contiene papeletas con importes MN o USD.")
    return output, errors


def record_error(cursor: pyodbc.Cursor, content: bytes, filename: str, hash_value: str, metadata: dict[str, str], message: str) -> None:
    found = cursor.execute("SELECT id FROM dbo.efc_conc_analiticos_importaciones WHERE hash_archivo=?", hash_value).fetchone()
    if found:
        cursor.execute("UPDATE dbo.efc_conc_analiticos_importaciones SET estado='ERROR',mensaje_error=?,actualizado_en=GETDATE() WHERE id=?", message, found[0])
    else:
        cursor.execute("INSERT dbo.efc_conc_analiticos_importaciones(hash_archivo,mensaje_uid,correo_origen,asunto,fecha_recepcion,nombre_archivo,archivo,version_lector,estado,mensaje_error) VALUES(?,?,?,?,?,?,?,?, 'ERROR',?)", hash_value, metadata.get("uid"), metadata.get("from"), metadata.get("subject"), metadata.get("received"), filename, pyodbc.Binary(content), VERSION, message)


def import_attachment(connection: pyodbc.Connection, content: bytes, filename: str, metadata: dict[str, str]) -> dict:
    hash_value = hashlib.sha256(content).hexdigest()
    cursor = connection.cursor()
    found = cursor.execute("SELECT id,estado FROM dbo.efc_conc_analiticos_importaciones WHERE hash_archivo=?", hash_value).fetchone()
    if found and found[1] == "IMPORTADA":
        return {"duplicate": True, "id": int(found[0]), "papeletas": 0, "errors": 0}
    try:
        stations, aliases = catalog(cursor)
        papers, errors = parse_workbook(content, filename, stations, aliases)
    except Exception as exc:
        record_error(cursor, content, filename, hash_value, metadata, str(exc)); connection.commit(); raise
    try:
        if found:
            import_id = int(found[0])
            cursor.execute("DELETE FROM dbo.efc_conc_analiticos_errores WHERE importacion_id=?", import_id)
            cursor.execute("DELETE FROM dbo.efc_conc_analiticos_papeletas WHERE importacion_id=?", import_id)
            cursor.execute("UPDATE dbo.efc_conc_analiticos_importaciones SET mensaje_uid=?,correo_origen=?,asunto=?,fecha_recepcion=?,nombre_archivo=?,archivo=?,version_lector=?,estado='PROCESANDO',total_papeletas=0,total_errores=0,mensaje_error=NULL,actualizado_en=GETDATE() WHERE id=?", metadata.get("uid"), metadata.get("from"), metadata.get("subject"), metadata.get("received"), filename, pyodbc.Binary(content), VERSION, import_id)
        else:
            import_id = int(cursor.execute("INSERT dbo.efc_conc_analiticos_importaciones(hash_archivo,mensaje_uid,correo_origen,asunto,fecha_recepcion,nombre_archivo,archivo,version_lector,estado) OUTPUT INSERTED.id VALUES(?,?,?,?,?,?,?,?, 'PROCESANDO')", hash_value, metadata.get("uid"), metadata.get("from"), metadata.get("subject"), metadata.get("received"), filename, pyodbc.Binary(content), VERSION).fetchone()[0])
        cursor.fast_executemany = True
        if papers:
            cursor.executemany("INSERT dbo.efc_conc_analiticos_papeletas(importacion_id,hoja,fila_origen,fecha_reportada,fecha_original,hora_original,estacion_original,estacion_nombre_original,estacion_id,estado_estacion,cuenta_mn_original,cuenta_mn_normalizada,remesa_numero,dice_contener_mn,real_mn,diferencia_mn,dice_contener_usd,real_usd,diferencia_usd,datos_originales) VALUES(?, 'PLANILLA',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", [(import_id, *paper) for paper in papers])
        if errors:
            cursor.executemany("INSERT dbo.efc_conc_analiticos_errores(importacion_id,hoja,fila_origen,tipo,detalle,datos_originales) VALUES(?, 'PLANILLA',?,?,?,?)", [(import_id, *error) for error in errors])
        cursor.execute("UPDATE dbo.efc_conc_analiticos_importaciones SET estado='IMPORTADA',total_papeletas=?,total_errores=?,actualizado_en=GETDATE() WHERE id=?", len(papers), len(errors), import_id)
        connection.commit()
        return {"duplicate": False, "id": import_id, "papeletas": len(papers), "errors": len(errors)}
    except Exception:
        connection.rollback()
        raise


def sync() -> dict:
    host = os.environ.get("EFC_CONC_ANALITICOS_IMAP_HOST", "imap.gmail.com")
    port = int(os.environ.get("EFC_CONC_ANALITICOS_IMAP_PORT", "993"))
    folder = os.environ.get("EFC_CONC_ANALITICOS_IMAP_FOLDER", "INBOX")
    user, password = os.environ.get("EFC_CONC_ANALITICOS_MAIL_USER", "").strip(), os.environ.get("EFC_CONC_ANALITICOS_MAIL_PASSWORD", "")
    if not user or not password:
        raise RuntimeError("Faltan EFC_CONC_ANALITICOS_MAIL_USER y EFC_CONC_ANALITICOS_MAIL_PASSWORD.")
    totals = {"messages": 0, "attachments": 0, "imported": 0, "duplicates": 0, "errors": 0}
    with db_connection() as connection, imaplib.IMAP4_SSL(host, port) as mailbox:
        ensure_schema(connection.cursor()); connection.commit()
        mailbox.login(user, password)
        if mailbox.select(folder, readonly=True)[0] != "OK": raise RuntimeError("No fue posible abrir el buzón de Analíticos.")
        status, data = mailbox.search(None, "SUBJECT", "ANALITICOS")
        if status != "OK": raise RuntimeError("No fue posible buscar correos de Analíticos.")
        for sequence in data[0].split():
            status, payload = mailbox.fetch(sequence, "(UID RFC822)")
            if status != "OK" or not payload or not payload[0]: totals["errors"] += 1; continue
            header, raw = payload[0]
            uid = header.decode(errors="ignore").split("UID ", 1)[-1].split(")", 1)[0].strip()
            message = email.message_from_bytes(raw); subject = decode(message.get("Subject"))
            if "ANALITICOS" not in subject.upper(): continue
            totals["messages"] += 1
            received = parsedate_to_datetime(message.get("Date")).replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S") if message.get("Date") else None
            metadata = {"uid": uid, "from": decode(message.get("From")), "subject": subject, "received": received}
            for part in message.walk():
                filename = decode(part.get_filename())
                if not filename or not filename.lower().endswith((".xls", ".xlsx")): continue
                content = part.get_payload(decode=True) or b""
                if not content: totals["errors"] += 1; continue
                totals["attachments"] += 1
                try:
                    result = import_attachment(connection, content, filename, metadata)
                    totals["duplicates" if result["duplicate"] else "imported"] += 1
                except Exception as exc:
                    connection.rollback(); totals["errors"] += 1; print(f"Adjunto {filename}: {exc}", file=sys.stderr)
    return totals


def main() -> int:
    argparse.ArgumentParser(description="Importa Analíticos REGIO a TG.").parse_args()
    load_env_file()
    try:
        print(json.dumps(sync(), ensure_ascii=False)); return 0
    except Exception as exc:
        print(str(exc), file=sys.stderr); return 1


if __name__ == "__main__":
    raise SystemExit(main())
